from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum, Count
import openpyxl
from openpyxl.styles import Font
from .models import Materiel,Demande # Ajoutez Demande si nécessaire
from .forms import MaterielForm
from django.views.decorators.http import require_http_methods
from django.views.decorators.cache import never_cache
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login as auth_login
from django.urls import reverse, reverse_lazy

def liste_materiels(request):
    action = request.GET.get("action")
    id_materiel = request.GET.get("id")

    form = None
    edit_obj = None
    delete_obj = None

    if action == "add":
        form = MaterielForm()

    elif action == "edit" and id_materiel:
        edit_obj = get_object_or_404(Materiel, id=id_materiel)
        form = MaterielForm(instance=edit_obj)

    elif action == "delete" and id_materiel:
        delete_obj = get_object_or_404(Materiel, id=id_materiel)

    query = request.GET.get("q", "")
    materiels = Materiel.objects.all().order_by('-date_ajout')
    if query:
        materiels = materiels.filter(
            Q(nom__icontains=query) | Q(categorie__icontains=query)
        )

    # Statistiques publiques (globales)
    stats_qs = Materiel.objects.aggregate(
        total_quantite=Sum('quantite'),
        total_bon=Sum('quantite_bon'),
        total_mauvais=Sum('quantite_mauvais'),
        total_categories=Count('categorie', distinct=True),
    )

    context = {
        "materiels": materiels,
        "action": action,
        "form": form,
        "edit_obj": edit_obj,
        "delete_obj": delete_obj,
        "query": query,
        # stats
        'total_materiels': Materiel.objects.count(),
        'total_quantite': stats_qs.get('total_quantite') or 0,
        'total_bon': stats_qs.get('total_bon') or 0,
        'total_mauvais': stats_qs.get('total_mauvais') or 0,
        'total_categories': stats_qs.get('total_categories') or 0,
    }

    return render(request, "gestion/main.html", context)

def ajouter_materiel(request):
    if request.method == "POST":
        form = MaterielForm(request.POST)
        if form.is_valid():
            m = form.save(commit=False)
            total = m.quantite
            if m.quantite_mauvais == 0:
                m.quantite_bon = total
            m.save()
    # Rediriger vers l'URL de retour si fournie (préserve l'interface admin)
    next_url = request.POST.get('next') or request.GET.get('next')
    if next_url and str(next_url).startswith('/'):
        return redirect(next_url)
    return redirect("liste_materiels")

def modifier_materiel(request, pk):
    materiel = get_object_or_404(Materiel, id=pk)
    if request.method == "POST":
        form = MaterielForm(request.POST, instance=materiel)
        if form.is_valid():
            form.save()
    next_url = request.POST.get('next') or request.GET.get('next')
    if next_url and str(next_url).startswith('/'):
        return redirect(next_url)
    return redirect("liste_materiels")

def supprimer_materiel(request, pk):
    materiel = get_object_or_404(Materiel, id=pk)
    materiel.delete()
    next_url = request.POST.get('next') or request.GET.get('next')
    if next_url and str(next_url).startswith('/'):
        # Retirer les paramètres de query (?action=...&id=...) pour éviter de rouvrir un modal
        clean_next = str(next_url).split('?')[0]
        return redirect(clean_next)
    return redirect("liste_materiels")

def exporter_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matériels"

    headers = ['ID', 'Nom', 'Catégorie', 'Quantité totale', 'Bon', 'Mauvais', 'Date d\'ajout']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    materiels = Materiel.objects.all()
    for row_idx, m in enumerate(materiels, 2):
        ws.cell(row=row_idx, column=1, value=m.id)
        ws.cell(row=row_idx, column=2, value=m.nom)
        ws.cell(row=row_idx, column=3, value=m.categorie)
        ws.cell(row=row_idx, column=4, value=m.quantite)
        ws.cell(row=row_idx, column=5, value=m.quantite_bon)
        ws.cell(row=row_idx, column=6, value=m.quantite_mauvais)
        ws.cell(row=row_idx, column=7, value=m.date_ajout.strftime("%Y-%m-%d %H:%M"))

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="materiels.xlsx"'
    wb.save(response)
    return response

def custom_login(request):
    # Si l'utilisateur est déjà connecté, redirigez-le vers liste_demande
    if request.user.is_authenticated:
        return redirect('liste_demande')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, f'Bienvenue {username} !')
            return redirect('liste_demande')
        else:
            messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')
    
    return render(request, 'gestion/login.html')

def register_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        
        # Validation des données
        if password != confirm_password:
            messages.error(request, 'Les mots de passe ne correspondent pas')
            return render(request, 'gestion/login.html')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Ce nom d\'utilisateur existe déjà')
            return render(request, 'gestion/login.html')
        
        if User.objects.filter(email=email).exists():
            messages.error(request, 'Cet email est déjà utilisé')
            return render(request, 'gestion/login.html')
        
        # Créer l'utilisateur
        try:
            user = User.objects.create_user(
                username=username,
                password=password,
                email=email,
                first_name=first_name,
                last_name=last_name
            )
            messages.success(request, 'Compte créé avec succès ! Vous pouvez maintenant vous connecter.')
            return redirect('custom_login')  # Rediriger vers la page de connexion
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du compte: {str(e)}')
            return render(request, 'gestion/login.html')
    
    return render(request, 'gestion/login.html')

@login_required
@never_cache
@login_required(login_url=reverse_lazy('login'))
def liste_demande(request):
    # Votre logique pour afficher les demandes
    # Seuls les utilisateurs connectés peuvent accéder à cette vue
    try:
        # Montrer seulement les demandes du user connecté (les admins peuvent voir tout depuis le dashboard)
        if request.user.is_superuser:
            demandes = Demande.objects.all().order_by('-date_demande')
        else:
            demandes = Demande.objects.filter(utilisateur=request.user).order_by('-date_demande')
    except Exception:
        demandes = []

    # Matériels à afficher dans le tableau (ceux présents en base)
    materiels = Materiel.objects.all().order_by('-date_ajout')

    # Statistiques simples pour l'utilisateur
    materiels_disponibles = Materiel.objects.filter(quantite_bon__gt=0).count()
    # Total d'unités empruntées : nombre total de quantités approuvées dans les demandes
    materiels_empruntes = Demande.objects.filter(statut='approuvee').aggregate(total=Sum('quantite_demandee'))['total'] or 0

    context = {
        'user': request.user,
        'demandes': demandes,
        'materiels': materiels,
        'materiels_disponibles': materiels_disponibles,
        'materiels_empruntes': materiels_empruntes,
        'title': 'Liste des Demandes',
        'message': 'Bienvenue sur la page des demandes'
    }
    return render(request, 'gestion/liste_demande.html', context)


@require_http_methods(["POST"])

def logout_view(request):
    logout(request)
    messages.success(request, 'Vous avez été déconnecté avec succès. Veuillez vous reconnecter.')
    return redirect('login')


from django.utils.http import url_has_allowed_host_and_scheme  # Ajoutez cette importation

def admin_login(request):
    # Si l'utilisateur est déjà connecté et est admin, rediriger vers le dashboard
    if request.user.is_authenticated and request.user.is_superuser:
        return redirect('admin_dashboard')
    
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            auth_login(request, user)
        # Vérifier si l'utilisateur a des demandes dont le statut a changé et qui ne sont pas encore notifiées
        to_notify = Demande.objects.filter(utilisateur=request.user, user_notified=False).exclude(statut='en_attente')
        for d in to_notify:
            if d.statut == 'approuvee':
                messages.success(request, f"Votre demande pour '{d.materiel.nom}' ({d.quantite_demandee}) a été approuvée.")
            elif d.statut == 'refusee':
                messages.error(request, f"Votre demande pour '{d.materiel.nom}' ({d.quantite_demandee}) a été rejetée.")
            # Marquer comme notifié
            d.user_notified = True
            d.save()

            
            # Redirection personnalisée pour les admins
            if user.is_superuser:
                next_url = request.POST.get('next') or request.GET.get('next')
                if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts=None):
                    return redirect(next_url)
                return redirect('admin_dashboard')  # Rediriger vers le dashboard admin
            else:
                # Si un utilisateur non-admin tente de se connecter via admin_login
                messages.error(request, "Accès réservé aux administrateurs")
                return redirect('liste_materiels')
        else:
            messages.error(request, "Identifiants invalides pour l'accès administrateur")
    
    else:
        form = AuthenticationForm()
    
    return render(request, 'gestion/admin_login.html', {'form': form})

@login_required(login_url=reverse_lazy('login'))
def admin_dashboard(request):
    if not request.user.is_superuser:
        messages.error(request, "Accès non autorisé")
        return redirect('liste_materiels')
    
    # Statistiques pour le dashboard admin
    total_materiels = Materiel.objects.count()
    total_demandes = Demande.objects.count()
    total_users = User.objects.count()

    # Comptes actifs / staff / superusers
    active_users = User.objects.filter(is_active=True).count()
    staff_users = User.objects.filter(is_staff=True).count()
    superusers = User.objects.filter(is_superuser=True).count()

    # Récupérer les matériels (utilisés dans le tableau)
    query = request.GET.get('q', '')
    materiels = Materiel.objects.all().order_by('-date_ajout')
    if query:
        materiels = materiels.filter(Q(nom__icontains=query) | Q(categorie__icontains=query))

    # Gérer action/form pour modal add/edit/delete (comme dans liste_materiels)
    action = request.GET.get('action')
    id_materiel = request.GET.get('id')
    form = None
    edit_obj = None
    delete_obj = None
    if action == 'add':
        form = MaterielForm()
    elif action == 'edit' and id_materiel:
        # Récupération sûre : si l'objet n'existe plus, ne pas lever 404
        edit_obj = Materiel.objects.filter(id=id_materiel).first()
        if edit_obj:
            form = MaterielForm(instance=edit_obj)
        else:
            form = None
            messages.warning(request, "Le matériel demandé est introuvable ou a été supprimé.")
            action = None
            id_materiel = None
    elif action == 'delete' and id_materiel:
        delete_obj = Materiel.objects.filter(id=id_materiel).first()
        if not delete_obj:
            # Si l'objet a déjà été supprimé, neutraliser l'action pour éviter une 404
            action = None
            id_materiel = None

    # Regrouper les demandes par utilisateur pour l'affichage
    demandes = Demande.objects.select_related('utilisateur').all().order_by('-date_demande')
    demandes_par_utilisateur = {}
    for d in demandes:
        user_key = d.utilisateur
        demandes_par_utilisateur.setdefault(user_key, []).append(d)

    # Liste complète des demandes (pour tableau administrateur)
    all_demandes = Demande.objects.select_related('utilisateur', 'materiel').all().order_by('-date_demande')

    demandes_en_attente = Demande.objects.filter(statut='en_attente').count()
    demandes_approuvees = Demande.objects.filter(statut='approuvee').count()
    demandes_rejetees = Demande.objects.filter(statut='refusee').count()

    context = {
        'total_materiels': total_materiels,
        'total_demandes': total_demandes,
        'total_users': total_users,
        'active_users': active_users,
        'staff_users': staff_users,
        'superusers': superusers,
        'materiels': materiels,
        'query': query,
        'action': action,
        'form': form,
        'edit_obj': edit_obj,
        'delete_obj': delete_obj,
        'demandes_par_utilisateur': demandes_par_utilisateur,
        'all_demandes': all_demandes,
        'demandes_en_attente': demandes_en_attente,
        'demandes_approuvees': demandes_approuvees,
        'demandes_rejetees': demandes_rejetees,
        'dernieres_demandes': demandes[:5],
        'derniers_materiels': materiels[:5],
    }

    return render(request, 'gestion/admin.html', context)


@login_required(login_url=reverse_lazy('login'))
def emprunter_materiel(request, pk):
    """Créer une demande d'emprunt pour l'utilisateur connecté (quantité 1 par défaut).
    Le statut initial est 'en_attente'.
    """
    materiel = get_object_or_404(Materiel, id=pk)

    # Vérifier disponibilité
    if materiel.quantite_bon <= 0:
        messages.error(request, "Désolé, ce matériel est actuellement indisponible.")
        return redirect('liste_demande')

    # Créer la demande (quantité 1 par défaut)
    try:
        Demande.objects.create(
            utilisateur=request.user,
            materiel=materiel,
            quantite_demandee=1,
            statut='en_attente',
            nom_demandeur=(request.user.get_full_name() or request.user.username),
            email_demandeur=request.user.email
        )
        messages.success(request, f"Demande d'emprunt pour '{materiel.nom}' enregistrée.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la création de la demande: {e}")

    # Rediriger vers la section 'Mes demandes' pour que l'utilisateur voie immédiatement sa demande
    try:
        return redirect(reverse('liste_demande') + '#mes-demandes')
    except Exception:
        return redirect('liste_demande')


@login_required(login_url=reverse_lazy('login'))
@require_http_methods(["POST"])
def annuler_demande(request, pk):
    """Permet à l'utilisateur d'annuler sa propre demande si elle est en attente.
    Suppression définitive de la demande (option simple).
    """
    # Récupérer la demande en s'assurant que c'est bien le propriétaire
    demande = get_object_or_404(Demande, id=pk, utilisateur=request.user)

    # Vérifier statut
    if demande.statut != 'en_attente':
        messages.error(request, "Seules les demandes en attente peuvent être annulées.")
        try:
            return redirect(reverse('liste_demande') + '#mes-demandes')
        except Exception:
            return redirect('liste_demande')

    # Supprimer la demande
    demande.delete()
    messages.success(request, "Votre demande a été annulée.")
    try:
        return redirect(reverse('liste_demande') + '#mes-demandes')
    except Exception:
        return redirect('liste_demande')


@login_required(login_url=reverse_lazy('login'))
def creer_demande(request):
    """Créer une demande depuis le formulaire modal sur la page demandes."""
    if request.method != 'POST':
        messages.error(request, "Méthode non autorisée")
        return redirect('liste_demande')

    materiel_id = request.POST.get('materiel_id')
    quantite = request.POST.get('quantite')
    nom = request.POST.get('nom')
    email = request.POST.get('email')
    raison = request.POST.get('raison', '')

    if not materiel_id or not quantite:
        messages.error(request, "Données incomplètes pour la demande.")
        return redirect('liste_demande')

    try:
        materiel = Materiel.objects.get(id=int(materiel_id))
        qte = int(quantite)
    except Exception:
        messages.error(request, "Matériel ou quantité invalide.")
        return redirect('liste_demande')

    if qte <= 0:
        messages.error(request, "La quantité doit être au moins 1.")
        return redirect('liste_demande')

    if materiel.quantite_bon < qte:
        messages.error(request, "Quantité demandée supérieure à la quantité disponible en bon état.")
        return redirect('liste_demande')

    # Utilisateur connecté sera le propriétaire ; nom/email sont des champs informatifs
    try:
        Demande.objects.create(
            utilisateur=request.user,
            materiel=materiel,
            quantite_demandee=qte,
            statut='en_attente',
            raison=raison,
            nom_demandeur=nom or (request.user.get_full_name() or request.user.username),
            email_demandeur=email or request.user.email,
        )
        messages.success(request, "Demande envoyée à l'administrateur.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la création de la demande: {e}")

    try:
        return redirect(reverse('liste_demande') + '#mes-demandes')
    except Exception:
        return redirect('liste_demande')



@login_required(login_url=reverse_lazy('login'))
def approuver_demande(request, pk):
    """Vue pour approuver une demande (accessible aux superusers/admins).
    Lors de l'approbation, on diminue la quantité totale du matériel si disponible.
    """
    if not request.user.is_superuser:
        messages.error(request, "Accès non autorisé")
        return redirect('liste_materiels')

    demande = get_object_or_404(Demande, id=pk)
    materiel = demande.materiel
    qte = demande.quantite_demandee

    if demande.statut == 'approuvee':
        messages.info(request, "Cette demande est déjà approuvée.")
        return redirect('admin_dashboard')

    if materiel.quantite_bon < qte:
        messages.error(request, "Quantité insuffisante pour approuver la demande.")
        return redirect('admin_dashboard')

    # Réduire la quantité totale (le save() du modèle réajustera quantite_bon)
    materiel.quantite = materiel.quantite - qte
    if materiel.quantite < 0:
        materiel.quantite = 0
    materiel.save()

    demande.statut = 'approuvee'
    # Marquer pour notification utilisateur
    demande.user_notified = False
    demande.save()
    messages.success(request, f"Demande #{demande.id} approuvée.")
    # Rediriger vers le bloc utilisateur concerné pour mise à jour visuelle
    try:
        return redirect(reverse('admin_dashboard') + f'#user-{demande.utilisateur.id}')
    except Exception:
        return redirect('admin_dashboard')


@login_required(login_url=reverse_lazy('login'))
def rejeter_demande(request, pk):
    """Vue pour rejeter une demande (accessible aux superusers/admins)."""
    if not request.user.is_superuser:
        messages.error(request, "Accès non autorisé")
        return redirect('liste_materiels')

    demande = get_object_or_404(Demande, id=pk)
    if demande.statut == 'refusee':
        messages.info(request, "Cette demande est déjà rejetée.")
        return redirect('admin_dashboard')

    demande.statut = 'refusee'
    demande.user_notified = False
    demande.save()
    messages.success(request, f"Demande #{demande.id} rejetée.")
    # Rediriger vers le bloc utilisateur concerné pour mise à jour visuelle
    try:
        return redirect(reverse('admin_dashboard') + f'#user-{demande.utilisateur.id}')
    except Exception:
        return redirect('admin_dashboard')


@login_required(login_url=reverse_lazy('login'))
def detail_demande(request, pk):
    """Afficher les détails d'une demande (accessible aux admins et au propriétaire)."""
    demande = get_object_or_404(Demande, id=pk)

    # Autoriser l'accès au propriétaire ou aux admins
    if not (request.user.is_superuser or demande.utilisateur == request.user):
        messages.error(request, "Accès non autorisé à cette demande.")
        return redirect('liste_materiels')

    return render(request, 'gestion/detail_demande.html', {'demande': demande})



def logout_admin(request):
    """Déconnexion spécifique pour l'admin avec message de confirmation"""
    if request.user.is_authenticated:
        username = request.user.username
        logout(request)
        messages.success(request, f"Déconnexion réussie. À bientôt {username}!")
    
    return redirect('admin_login')